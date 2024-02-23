from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles.colors import BLUE
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def format_file(ws):
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    header_font = Font(bold=True, color='FFFFFF', size=14)
    header_fill = PatternFill(start_color='4A86E8', end_color='4A86E8', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    cell_font = Font(bold=False, color='000000', size=12)
    cell_alignment = Alignment(horizontal='left', vertical='center')

    link_font = Font(color=BLUE, underline='single', size=12)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 26.25

    for col in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            if cell.row == 1:       # Header row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            else:                   # Data rows
                cell.alignment = cell_alignment
                if 5 <= cell.column <= 6:
                    cell.font = link_font
                    cell.value = f'=HYPERLINK("{cell.value}", "{cell.value}")'
                else:
                    cell.font = cell_font
